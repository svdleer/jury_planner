from io import BytesIO
import csv
from datetime import datetime
from typing import List, Dict, Any
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd

class BaseExporter:
    """Base class for all exporters"""
    
    def __init__(self):
        self.title = "Water Polo Jury Schedule"
        self.headers = [
            'Date', 'Time', 'Home Team', 'Away Team', 
            'Location', 'Competition', 'Jury Team', 'Duty'
        ]
    
    def export(self, data: List[Dict[str, Any]]) -> BytesIO:
        """Export data to specified format"""
        raise NotImplementedError()

class PDFExporter(BaseExporter):
    """Export schedule to PDF format"""
    
    def export(self, data: List[Dict[str, Any]]) -> BytesIO:
        """Export data to PDF"""
        buffer = BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        
        # Create custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Build content
        content = []
        
        # Title
        title = Paragraph(self.title, title_style)
        content.append(title)
        content.append(Spacer(1, 20))
        
        # Export date
        export_date = Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal'])
        content.append(export_date)
        content.append(Spacer(1, 20))
        
        if not data:
            no_data = Paragraph("No schedule data available for the selected criteria.", styles['Normal'])
            content.append(no_data)
        else:
            # Prepare table data
            table_data = [self.headers]
            
            for row in data:
                table_row = [
                    row['date'].strftime('%Y-%m-%d') if hasattr(row['date'], 'strftime') else str(row['date']),
                    row['time'].strftime('%H:%M') if hasattr(row['time'], 'strftime') else str(row['time']),
                    str(row['home_team']),
                    str(row['away_team']),
                    str(row['location'] or ''),
                    str(row['competition'] or ''),
                    str(row['jury_team']),
                    str(row['duty']).replace('_', ' ').title()
                ]
                table_data.append(table_row)
            
            # Create table
            table = Table(table_data)
            
            # Style the table
            table.setStyle(TableStyle([
                # Header row
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Data rows
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            
            content.append(table)
        
        # Build PDF
        doc.build(content)
        buffer.seek(0)
        
        return buffer

class ExcelExporter(BaseExporter):
    """Export schedule to Excel format"""
    
    def export(self, data: List[Dict[str, Any]]) -> BytesIO:
        """Export data to Excel"""
        buffer = BytesIO()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jury Schedule"
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add title
        ws['A1'] = self.title
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = center_alignment
        
        # Add export date
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws.merge_cells('A2:H2')
        
        # Add headers
        for col, header in enumerate(self.headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Add data
        if data:
            for row_idx, row_data in enumerate(data, 5):
                ws.cell(row=row_idx, column=1, value=row_data['date'].strftime('%Y-%m-%d') if hasattr(row_data['date'], 'strftime') else str(row_data['date']))
                ws.cell(row=row_idx, column=2, value=row_data['time'].strftime('%H:%M') if hasattr(row_data['time'], 'strftime') else str(row_data['time']))
                ws.cell(row=row_idx, column=3, value=str(row_data['home_team']))
                ws.cell(row=row_idx, column=4, value=str(row_data['away_team']))
                ws.cell(row=row_idx, column=5, value=str(row_data['location'] or ''))
                ws.cell(row=row_idx, column=6, value=str(row_data['competition'] or ''))
                ws.cell(row=row_idx, column=7, value=str(row_data['jury_team']))
                ws.cell(row=row_idx, column=8, value=str(row_data['duty']).replace('_', ' ').title())
                
                # Center align data
                for col in range(1, 9):
                    ws.cell(row=row_idx, column=col).alignment = center_alignment
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer

class CSVExporter(BaseExporter):
    """Export schedule to CSV format"""
    
    def export(self, data: List[Dict[str, Any]]) -> BytesIO:
        """Export data to CSV"""
        buffer = BytesIO()
        
        # Create CSV writer (using text mode wrapper)
        import io
        text_buffer = io.StringIO()
        writer = csv.writer(text_buffer)
        
        # Write headers
        writer.writerow(self.headers)
        
        # Write data
        for row in data:
            csv_row = [
                row['date'].strftime('%Y-%m-%d') if hasattr(row['date'], 'strftime') else str(row['date']),
                row['time'].strftime('%H:%M') if hasattr(row['time'], 'strftime') else str(row['time']),
                str(row['home_team']),
                str(row['away_team']),
                str(row['location'] or ''),
                str(row['competition'] or ''),
                str(row['jury_team']),
                str(row['duty']).replace('_', ' ').title()
            ]
            writer.writerow(csv_row)
        
        # Convert to bytes
        buffer = BytesIO(text_buffer.getvalue().encode('utf-8'))
        buffer.seek(0)
        
        return buffer

class TXTExporter(BaseExporter):
    """Export schedule to plain text format"""
    
    def export(self, data: List[Dict[str, Any]]) -> BytesIO:
        """Export data to plain text"""
        buffer = BytesIO()
        
        lines = []
        
        # Title and header
        lines.append(self.title)
        lines.append("=" * len(self.title))
        lines.append("")
        lines.append(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        lines.append("")
        
        if not data:
            lines.append("No schedule data available for the selected criteria.")
        else:
            # Calculate column widths
            col_widths = [len(header) for header in self.headers]
            
            for row in data:
                row_data = [
                    row['date'].strftime('%Y-%m-%d') if hasattr(row['date'], 'strftime') else str(row['date']),
                    row['time'].strftime('%H:%M') if hasattr(row['time'], 'strftime') else str(row['time']),
                    str(row['home_team']),
                    str(row['away_team']),
                    str(row['location'] or ''),
                    str(row['competition'] or ''),
                    str(row['jury_team']),
                    str(row['duty']).replace('_', ' ').title()
                ]
                
                for i, value in enumerate(row_data):
                    col_widths[i] = max(col_widths[i], len(value))
            
            # Add padding
            col_widths = [w + 2 for w in col_widths]
            
            # Header
            header_line = ""
            separator_line = ""
            for i, (header, width) in enumerate(zip(self.headers, col_widths)):
                header_line += header.ljust(width)
                separator_line += "-" * width
            
            lines.append(header_line.rstrip())
            lines.append(separator_line.rstrip())
            
            # Data rows
            for row in data:
                row_data = [
                    row['date'].strftime('%Y-%m-%d') if hasattr(row['date'], 'strftime') else str(row['date']),
                    row['time'].strftime('%H:%M') if hasattr(row['time'], 'strftime') else str(row['time']),
                    str(row['home_team']),
                    str(row['away_team']),
                    str(row['location'] or ''),
                    str(row['competition'] or ''),
                    str(row['jury_team']),
                    str(row['duty']).replace('_', ' ').title()
                ]
                
                data_line = ""
                for value, width in zip(row_data, col_widths):
                    data_line += value.ljust(width)
                
                lines.append(data_line.rstrip())
        
        # Write to buffer
        content = "\n".join(lines)
        buffer = BytesIO(content.encode('utf-8'))
        buffer.seek(0)
        
        return buffer
